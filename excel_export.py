from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from database import get_application

EXCEL_PATH = Path("applications.xlsx")

HEADERS = [
    "Время подачи",
    "Имя",
    "Дата рождения",
    "Город и страна",
    "Телефон",
    "Помещение без посторонних",
    "Устройства",
    "Модель устройства",
    "Наушники",
    "Время работы",
    "Опыт",
    "Telegram",
    "Источник",
    "Статус",
    "ID",
]

OLD_HEADERS = [
    "Время подачи",
    "Имя",
    "Дата рождения",
    "Город и страна",
    "Телефон",
    "Помещение без посторонних",
    "Устройства",
    "Модель устройства",
    "Наушники",
    "Время работы",
    "Опыт",
    "Telegram",
    "Статус",
    "ID",
]

def _init_sheet(ws):
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(HEADERS))}1"

def _ensure_headers(ws):
    if ws.max_row < 1:
        _init_sheet(ws)
        return
    current = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    if current[:len(OLD_HEADERS)] == OLD_HEADERS:
        insert_at = OLD_HEADERS.index("Статус") + 1
        ws.insert_cols(insert_at)
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(HEADERS))}1"

def _fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

def _format_submit_time(ts: str | None) -> str:
    if not ts:
        return datetime.now().strftime("%d.%m.%Y %H:%M")
    try:
        dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
        return dt.astimezone().strftime("%d.%m.%Y %H:%M")
    except Exception:
        return ts

def _format_source(value: str | None) -> str:
    if not value:
        return ""
    return "Сайт" if value.lower() == "site" else "Бот"

def append_application_row(data: dict[str, Any], user_id: int, status: str):
    app = get_application(user_id) or {}
    ts = _format_submit_time(app.get("last_apply_at") or app.get("created_at"))
    source = _format_source(app.get("source"))
    row = [
        ts,
        data.get("name", ""),
        data.get("age", ""),
        data.get("city", ""),
        data.get("phone", ""),
        data.get("living", ""),
        data.get("devices", ""),
        data.get("device_model", ""),
        data.get("headphones", ""),
        data.get("work_time", ""),
        data.get("experience", ""),
        data.get("telegram", ""),
        source,
        status,
        str(user_id),
    ]

    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        _ensure_headers(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки"
        _init_sheet(ws)

    ws.append(row)
    _fit_columns(ws)
    wb.save(EXCEL_PATH)

def update_application_status(user_id: int, status: str):
    if not EXCEL_PATH.exists():
        return
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    target_row = None
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=len(HEADERS)).value
        val_alt = ws.cell(row=row, column=len(HEADERS) - 1).value
        if str(val) == str(user_id) or str(val_alt) == str(user_id):
            target_row = row
    if not target_row:
        return
    ws.cell(row=target_row, column=HEADERS.index("Статус") + 1).value = status
    wb.save(EXCEL_PATH)
